from django.contrib import admin
from .models import Song, Artist
import sqlite3
from openpyxl import load_workbook, Workbook
from django.http import HttpResponse
import datetime


class ArtistAdmin(admin.ModelAdmin):

    def export_musics_from_artist(self, request, queryset) -> HttpResponse:

        # definindo configurações do arquivo a ser enviado 
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=artist_music.xlsx'

        # definindo arquivo .xlsx
        wb = Workbook()
        sheet = wb.active

        qtd = len(queryset)
        init = 1
        for artist in queryset:
            #pegando o artista e as músicas dele pelo método que já estava pronto no model
            artist = Artist.objects.get(pk=artist.id)
            musics = artist.get_songs()

            #renomeando a planilha do momento
            sheet.title = artist.name
          
            #setando os cabeçalhos
            sheet['B2'] = "TITLE"
            sheet['C2'] = "DURATION/MINUTES"
            sheet['D2'] = "NO SPOTIFY"
            sheet['E2'] = "NO YOUTUBE"

            row = 3
            # adicionando músicas do banco na planilha
            for music in musics:
                song = musics[music]

                sheet['B'+str(row)] = song['title']
                sheet['C'+str(row)] = str(datetime.timedelta(seconds=song['duration']))
                sheet['D'+str(row)] = "Sim" if song['spotify'] == 1 else "Não"
                sheet['E'+str(row)] = "Sim" if song['youtube'] == 1 else "Não"

                row += 1

            if init == qtd:
                break
            init += 1

            #sobrescrevendo a planilha
            sheet = wb.create_sheet()
        
        #os dados que estão na planilha são gravados no response e retornados
        wb.save(response)
        
        return response
        
    actions = [export_musics_from_artist]


admin.site.register(Song)
admin.site.register(Artist, ArtistAdmin)